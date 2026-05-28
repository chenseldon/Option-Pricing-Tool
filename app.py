"""
app.py - Flask Web Application Backend (Enhanced)

Complete OTC option pricing tool backend with full risk analytics.

Endpoints:
    GET  /                      : Serve the web interface (index.html)
    POST /api/calculate         : BSM pricing + Monte Carlo + Greeks + chart data
    POST /api/iv                : Implied volatility solver
    POST /api/portfolio         : Portfolio exposure aggregation (MTM + margin)
    POST /api/export            : Export results to Excel (.xlsx)
    POST /api/vol_smile         : Volatility smile from market prices
    POST /api/mc_distribution   : Monte Carlo payoff distribution histogram
    POST /api/sensitivity       : Sensitivity / stress test matrix

Usage:
    python app.py
    Then open: http://127.0.0.1:5000

Dependencies:
    pip install flask py_vollib numpy pandas matplotlib openpyxl

Author: GitHub Portfolio Project
"""

import os
import sys
import numpy as np
from io import BytesIO
from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Ensure sibling modules (bs_pricing, greeks_calc) are importable
sys.path.insert(0, os.path.dirname(__file__))

from bs_pricing import bs_price, monte_carlo_price, calc_implied_volatility, validate_params
from greeks_calc import calc_delta, calc_gamma, calc_vega, calc_theta, calc_rho

app = Flask(__name__)


# ── Shared helper: compute all 5 Greeks at one point ──────────────────────────

def _greeks_at(flag, S, K, T, r, sigma, q):
    return {
        "Delta": round(float(calc_delta(flag, S, K, T, r, sigma, q)), 6),
        "Gamma": round(float(calc_gamma(flag, S, K, T, r, sigma, q)), 6),
        "Vega":  round(float(calc_vega (flag, S, K, T, r, sigma, q)), 6),
        "Theta": round(float(calc_theta(flag, S, K, T, r, sigma, q)), 6),
        "Rho":   round(float(calc_rho  (flag, S, K, T, r, sigma, q)), 6),
    }


# ─────────────────────────────────────────────────────────
# Frontend Route
# ─────────────────────────────────────────────────────────

@app.route("/")
def index():
    """Serve the main web interface page."""
    return render_template("index.html")


# ─────────────────────────────────────────────────────────
# API: Full Calculation (Price + Greeks + Chart Data)
# ─────────────────────────────────────────────────────────

@app.route("/api/calculate", methods=["POST"])
def calculate():
    """BSM + Monte Carlo pricing, all Greeks, and chart curve data."""
    try:
        data  = request.get_json(force=True)
        flag  = str(data.get("flag", "c")).lower()
        S     = float(data["S"])
        K     = float(data["K"])
        T     = float(data["T"])
        r     = float(data["r"])
        sigma = float(data["sigma"])
        q     = float(data.get("q", 0.0))
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": f"Invalid input parameters: {e}"}), 400

    if flag not in ("c", "p"):
        return jsonify({"error": "flag must be 'c' (Call) or 'p' (Put)"}), 400

    valid, msg = validate_params(S, K, T, r, sigma, q)
    if not valid:
        return jsonify({"error": msg}), 400

    bsm_price_val = bs_price(flag, S, K, T, r, sigma, q)
    mc_price_val  = monte_carlo_price(flag, S, K, T, r, sigma, q, n_simulations=100000)
    greeks        = _greeks_at(flag, S, K, T, r, sigma, q)

    # Chart: 100 S-values spanning 0.6K – 1.4K
    S_arr = np.linspace(0.6 * K, 1.4 * K, 100)

    prices_arr    = [round(float(bs_price(flag, s, K, T, r, sigma, q)), 4) for s in S_arr]
    intrinsic_arr = [round(float(max(s - K, 0) if flag == "c" else max(K - s, 0)), 4)
                     for s in S_arr]
    greek_curves  = {
        "Delta": [round(float(calc_delta(flag, s, K, T, r, sigma, q)), 6) for s in S_arr],
        "Gamma": [round(float(calc_gamma(flag, s, K, T, r, sigma, q)), 6) for s in S_arr],
        "Vega":  [round(float(calc_vega (flag, s, K, T, r, sigma, q)), 6) for s in S_arr],
        "Theta": [round(float(calc_theta(flag, s, K, T, r, sigma, q)), 6) for s in S_arr],
        "Rho":   [round(float(calc_rho  (flag, s, K, T, r, sigma, q)), 6) for s in S_arr],
    }

    return jsonify({
        "bsm_price": round(float(bsm_price_val), 4),
        "mc_price":  round(float(mc_price_val),  4),
        "greeks": {k.lower(): v for k, v in greeks.items()},
        "price_curve": {
            "spots":    [round(float(s), 2) for s in S_arr],
            "prices":   prices_arr,
            "intrinsic": intrinsic_arr,
        },
        "greeks_curve": {k.lower(): v for k, v in greek_curves.items()},
        "K": K,
    })


# ─────────────────────────────────────────────────────────
# API: Implied Volatility Solver
# ─────────────────────────────────────────────────────────

@app.route("/api/iv", methods=["POST"])
def implied_vol():
    """Solve for implied volatility from a market price."""
    try:
        data         = request.get_json(force=True)
        flag         = str(data.get("flag", "c")).lower()
        market_price = float(data["market_price"])
        S            = float(data["S"])
        K            = float(data["K"])
        T            = float(data["T"])
        r            = float(data["r"])
        q            = float(data.get("q", 0.0))
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": f"Invalid input parameters: {e}"}), 400

    try:
        iv = calc_implied_volatility(flag, market_price, S, K, T, r, q)
        return jsonify({"implied_volatility": round(float(iv), 6)})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


# ─────────────────────────────────────────────────────────
# API: Portfolio Exposure Aggregation
# ─────────────────────────────────────────────────────────

@app.route("/api/portfolio", methods=["POST"])
def portfolio():
    """
    Compute per-position Greeks, MTM, P&L, and margin for a portfolio,
    plus aggregate net Greeks across all positions.

    Each position in request JSON:
        flag (str), S/K/T/r/sigma/q (float), qty (float, signed),
        entry_price (float), label (str), id (str)

    Margin formula (simplified):
        Long  : |qty| * entry_price   (premium paid)
        Short : |qty| * S * 0.15      (15% of underlying, simplified)
    """
    try:
        data      = request.get_json(force=True)
        positions = data.get("positions", [])
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    results = []
    totals  = {"net_delta": 0.0, "net_gamma": 0.0, "net_vega": 0.0,
                "net_theta": 0.0, "net_rho": 0.0,
                "total_mtm": 0.0, "total_pnl": 0.0, "total_margin": 0.0}

    for pos in positions:
        try:
            flag  = str(pos["flag"]).lower()
            S     = float(pos["S"])
            K     = float(pos["K"])
            T     = float(pos["T"])
            r     = float(pos["r"])
            sigma = float(pos["sigma"])
            q     = float(pos.get("q", 0.0))
            qty   = float(pos.get("qty", 1.0))
            entry = float(pos.get("entry", pos.get("entry_price", 0.0)))

            price          = float(bs_price(flag, S, K, T, r, sigma, q))
            greeks_single  = _greeks_at(flag, S, K, T, r, sigma, q)

            mtm    = round(price * qty, 4)
            pnl    = round((price - entry) * qty, 4)
            margin = round(abs(qty) * entry, 4) if qty > 0 \
                     else round(abs(qty) * S * 0.15, 4)

            result_pos = {
                "id":        pos.get("id", ""),
                "label":     pos.get("label", ""),
                "flag":      flag, "S": S, "K": K,
                "T":         round(T, 6),
                "T_days":    round(T * 365),
                "r":         r, "sigma": sigma, "q": q,
                "qty":       qty,
                "entry":     round(entry, 4),
                "mtm":       mtm, "pnl": pnl, "margin": margin,
                "delta_net": round(greeks_single["Delta"] * qty, 6),
                "gamma_net": round(greeks_single["Gamma"] * qty, 6),
                "vega_net":  round(greeks_single["Vega"]  * qty, 6),
                "theta_net": round(greeks_single["Theta"] * qty, 6),
                "rho_net":   round(greeks_single["Rho"]   * qty, 6),
                "error":     None,
            }

            totals["net_delta"]    = round(totals["net_delta"]    + result_pos["delta_net"], 6)
            totals["net_gamma"]    = round(totals["net_gamma"]    + result_pos["gamma_net"], 6)
            totals["net_vega"]     = round(totals["net_vega"]     + result_pos["vega_net"],  6)
            totals["net_theta"]    = round(totals["net_theta"]    + result_pos["theta_net"], 6)
            totals["net_rho"]      = round(totals["net_rho"]      + result_pos["rho_net"],   6)
            totals["total_mtm"]    = round(totals["total_mtm"]    + mtm,    4)
            totals["total_pnl"]    = round(totals["total_pnl"]    + pnl,    4)
            totals["total_margin"] = round(totals["total_margin"] + margin, 4)

        except Exception as e:
            result_pos = {**pos, "error": str(e)}

        results.append(result_pos)

    return jsonify({"positions": results, "totals": totals})


# ─────────────────────────────────────────────────────────
# API: Vol Smile Preset Prices
# ─────────────────────────────────────────────────────────

@app.route("/api/smile_presets", methods=["POST"])
def smile_presets():
    """Return BSM prices for a list of strikes — used to pre-populate the Vol Smile input table."""
    try:
        data   = request.get_json(force=True)
        flag   = str(data.get("flag", "c")).lower()
        S      = float(data["S"])
        T      = float(data["T"])
        r      = float(data["r"])
        sigma  = float(data["sigma"])
        q      = float(data.get("q", 0.0))
        strikes = data.get("strikes", [])
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": f"Invalid parameters: {e}"}), 400

    presets = []
    for K_i in strikes:
        try:
            price = round(float(bs_price(flag, S, float(K_i), T, r, sigma, q)), 2)
            presets.append({"K": float(K_i), "price": price})
        except Exception:
            presets.append({"K": float(K_i), "price": 0})
    return jsonify({"presets": presets})

@app.route("/api/vol_smile", methods=["POST"])
def vol_smile():
    """
    Compute implied volatility for a set of (K, market_price) pairs,
    producing a volatility smile / skew curve.

    Request JSON: {flag, S, T, r, q, smile_data: [{K, market_price}]}
    Response JSON: {smile_points: [{K, iv (%), error}]}
    """
    try:
        data       = request.get_json(force=True)
        flag       = str(data.get("flag", "c")).lower()
        S          = float(data["S"])
        T          = float(data["T"])
        r          = float(data["r"])
        q          = float(data.get("q", 0.0))
        smile_data = data.get("pairs", data.get("smile_data", []))
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": f"Invalid parameters: {e}"}), 400

    points = []
    for item in smile_data:
        try:
            K_i  = float(item["K"])
            mp_i = float(item["market_price"])
            iv   = calc_implied_volatility(flag, mp_i, S, K_i, T, r, q)
            points.append({"K": K_i, "iv": round(float(iv) * 100, 4), "error": None})
        except Exception as e:
            points.append({"K": float(item.get("K", 0)), "iv": None, "error": str(e)})

    return jsonify({"results": points})


# ─────────────────────────────────────────────────────────
# API: Monte Carlo Payoff Distribution
# ─────────────────────────────────────────────────────────

@app.route("/api/mc_distribution", methods=["POST"])
def mc_distribution():
    """
    Run Monte Carlo simulation and return the discounted payoff distribution.
    Visualises the full probability distribution of option value at expiry.

    Request JSON: standard option params + n_simulations (int, max 200000)
    Response JSON: {bins, counts, mean, median, p5, p95, option_price, n_simulations}
    """
    try:
        data  = request.get_json(force=True)
        flag  = str(data.get("flag", "c")).lower()
        S     = float(data["S"])
        K     = float(data["K"])
        T     = float(data["T"])
        r     = float(data["r"])
        sigma = float(data["sigma"])
        q     = float(data.get("q", 0.0))
        n     = min(int(data.get("n_paths", data.get("n_simulations", 50000))), 200000)
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": f"Invalid parameters: {e}"}), 400

    valid, msg = validate_params(S, K, T, r, sigma, q)
    if not valid:
        return jsonify({"error": msg}), 400

    np.random.seed(42)
    Z   = np.random.standard_normal(n)
    S_T = S * np.exp((r - q - 0.5 * sigma**2) * T + sigma * np.sqrt(T) * Z)

    payoffs    = np.maximum(S_T - K, 0) if flag == "c" else np.maximum(K - S_T, 0)
    discounted = np.exp(-r * T) * payoffs

    counts, edges = np.histogram(discounted, bins=50)
    centers       = ((edges[:-1] + edges[1:]) / 2).tolist()

    return jsonify({
        "bins":   [round(x, 4) for x in centers],
        "counts": counts.tolist(),
        "stats": {
            "mean":       round(float(np.mean(discounted)),           4),
            "median":     round(float(np.median(discounted)),         4),
            "ci95_lower": round(float(np.percentile(discounted,  5)), 4),
            "ci95_upper": round(float(np.percentile(discounted, 95)), 4),
            "std":        round(float(np.std(discounted)),            4),
            "n_paths":    n,
        },
    })


# ─────────────────────────────────────────────────────────
# API: Sensitivity / Stress Test Matrix
# ─────────────────────────────────────────────────────────

@app.route("/api/sensitivity", methods=["POST"])
def sensitivity():
    """
    Generate 5x5 sensitivity matrices for option price and Delta under
    combinations of shifted volatility (+/-20%, +/-10%) and time (+/-60d, +/-30d).

    Request JSON: standard option params
    Response JSON: {sigma_labels, T_labels, price_matrix, delta_matrix, base_price}
    """
    try:
        data  = request.get_json(force=True)
        flag  = str(data.get("flag", "c")).lower()
        S     = float(data["S"])
        K     = float(data["K"])
        T     = float(data["T"])
        r     = float(data["r"])
        sigma = float(data["sigma"])
        q     = float(data.get("q", 0.0))
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": f"Invalid parameters: {e}"}), 400

    valid, msg = validate_params(S, K, T, r, sigma, q)
    if not valid:
        return jsonify({"error": msg}), 400

    sigma_offsets = [-0.20, -0.10, 0.0, +0.10, +0.20]
    T_offsets     = [-60/365, -30/365, 0.0, +30/365, +60/365]
    sigma_labels  = ["σ-20%", "σ-10%", "σ base", "σ+10%", "σ+20%"]
    T_labels      = ["T-60d", "T-30d", "T base", "T+30d", "T+60d"]

    price_matrix, delta_matrix = [], []
    for ds in sigma_offsets:
        price_row, delta_row = [], []
        for dt in T_offsets:
            s_i = max(sigma + ds, 0.005)
            t_i = max(T     + dt, 1 / 365)
            try:
                p = round(float(bs_price(flag, S, K, t_i, r, s_i, q)), 4)
                d = round(float(calc_delta(flag, S, K, t_i, r, s_i, q)), 4)
            except Exception:
                p, d = None, None
            price_row.append(p)
            delta_row.append(d)
        price_matrix.append(price_row)
        delta_matrix.append(delta_row)

    return jsonify({
        "sigma_labels": sigma_labels,
        "t_labels":     T_labels,
        "price_matrix": price_matrix,
        "delta_matrix": delta_matrix,
        "base_price":   round(float(bs_price(flag, S, K, T, r, sigma, q)), 4),
    })


# ─────────────────────────────────────────────────────────
# API: Excel Export
# ─────────────────────────────────────────────────────────

@app.route("/api/export", methods=["POST"])
def export_excel():
    """
    Generate and return an Excel report from the current pricing session.

    Request JSON: {params, result, portfolio}
    Returns: option_pricing_report.xlsx as file download
    """
    try:
        data      = request.get_json(force=True)
        params    = data.get("params",    {})
        result    = data.get("result",    {})
        portfolio = data.get("portfolio", {})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    wb = openpyxl.Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="0D1626")
    HDR_FONT  = Font(color="4E9EFF", bold=True, size=12)
    SUB_FILL  = PatternFill("solid", fgColor="0D3A6B")
    SUB_FONT  = Font(color="FFFFFF", bold=True)
    BOLD      = Font(bold=True)
    CENTER    = Alignment(horizontal="center")

    def hdr(ws, row, ncols, text, fill=HDR_FILL, font=HDR_FONT):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.fill, c.font, c.alignment = fill, font, CENTER

    def row_vals(ws, r, vals, bold=False, fill=None):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            if bold: c.font = BOLD
            if fill: c.fill = fill

    # ── Sheet 1: Pricing Summary ──────────────────────────
    ws1 = wb.active
    ws1.title = "Pricing Summary"
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 36

    r = 1
    hdr(ws1, r, 3, "OTC Option Pricing Summary Report"); r += 1
    hdr(ws1, r, 3, "Input Parameters", SUB_FILL, SUB_FONT); r += 1
    for label, val in [
        ("Option Type",    "Call" if params.get("flag") == "c" else "Put"),
        ("Underlying (S)", params.get("S", "")),
        ("Strike (K)",     params.get("K", "")),
        ("Maturity T (yr)",params.get("T", "")),
        ("Risk-Free r",    params.get("r", "")),
        ("Volatility σ",   params.get("sigma", "")),
        ("Div Yield q",    params.get("q", "")),
    ]:
        row_vals(ws1, r, [label, val]); r += 1

    r += 1
    hdr(ws1, r, 3, "Pricing Results", SUB_FILL, SUB_FONT); r += 1
    row_vals(ws1, r, ["BSM Price",         result.get("bsm_price", "")]); r += 1
    row_vals(ws1, r, ["Monte Carlo Price", result.get("mc_price",  "")]); r += 1

    r += 1
    hdr(ws1, r, 3, "Greeks at Current S", SUB_FILL, SUB_FONT); r += 1
    ws1.cell(row=r, column=1, value="Greek").font = BOLD
    ws1.cell(row=r, column=2, value="Value").font = BOLD
    ws1.cell(row=r, column=3, value="Description").font = BOLD
    r += 1
    greeks_data = result.get("greeks", {})
    for gname, gdesc in [
        ("Delta", "Price change per $1 move in underlying"),
        ("Gamma", "Delta change per $1 move in underlying"),
        ("Vega",  "Price change per 1% move in volatility"),
        ("Theta", "Daily time decay"),
        ("Rho",   "Price change per 1% move in rate"),
    ]:
        row_vals(ws1, r, [gname, greeks_data.get(gname, ""), gdesc]); r += 1

    # ── Sheet 2: Price Curve Data ─────────────────────────
    chart_data = result.get("chart", {})
    if chart_data:
        ws2 = wb.create_sheet("Price Curve Data")
        for ci, (col, w) in enumerate([("A",12),("B",14),("C",14),("D",14)], 0):
            ws2.column_dimensions[col].width = w
        hdr(ws2, 1, 4, "Option Price Curve Data")
        for ci, h in enumerate(["Underlying S","BSM Price","Intrinsic Value","Time Value"], 1):
            ws2.cell(row=2, column=ci, value=h).font = BOLD
        S_vals = chart_data.get("S_values", [])
        prices = chart_data.get("prices",   [])
        intr   = chart_data.get("intrinsic",[])
        for i, sv in enumerate(S_vals):
            p  = prices[i] if i < len(prices) else ""
            iv = intr[i]   if i < len(intr)   else ""
            tv = round(p - iv, 4) if isinstance(p, (int,float)) and isinstance(iv,(int,float)) else ""
            ws2.append([sv, p, iv, tv])

    # ── Sheet 3: Greeks Curves Data ───────────────────────
    gc = chart_data.get("greek_curves", {})
    if gc:
        ws3 = wb.create_sheet("Greeks Curves Data")
        headers = ["Underlying S","Delta","Gamma","Vega","Theta","Rho"]
        for ci, h in enumerate(headers, 1):
            col = openpyxl.utils.get_column_letter(ci)
            ws3.column_dimensions[col].width = 14
            ws3.cell(row=1, column=ci, value=h).font = BOLD
        for i, sv in enumerate(S_vals):
            row_data = [sv] + [gc[g][i] if i < len(gc.get(g, [])) else ""
                               for g in ["Delta","Gamma","Vega","Theta","Rho"]]
            ws3.append(row_data)

    # ── Sheet 4: Portfolio ────────────────────────────────
    positions = portfolio.get("positions", [])
    totals    = portfolio.get("totals",    {})
    if positions:
        ws4 = wb.create_sheet("Portfolio")
        ph  = ["#","Label","Type","S","K","T(yr)","σ","Qty","Entry",
               "BSM Price","MTM","P&L","Δ×Qty","Γ×Qty","V×Qty","Θ×Qty","ρ×Qty","Margin"]
        for ci, h in enumerate(ph, 1):
            col = openpyxl.utils.get_column_letter(ci)
            ws4.column_dimensions[col].width = 12
            ws4.cell(row=1, column=ci, value=h).font = BOLD
        for pi, pos in enumerate(positions, 1):
            wg = pos.get("w_greeks", {})
            ws4.append([
                pi, pos.get("label",""),
                "Call" if pos.get("flag")=="c" else "Put",
                pos.get("S",""), pos.get("K",""), pos.get("T",""),
                pos.get("sigma",""), pos.get("qty",""), pos.get("entry_price",""),
                pos.get("bsm_price",""), pos.get("mtm",""), pos.get("pnl",""),
                wg.get("Delta",""), wg.get("Gamma",""), wg.get("Vega",""),
                wg.get("Theta",""), wg.get("Rho",""), pos.get("margin",""),
            ])
        # Total row
        tr = len(positions) + 2
        for ci, val in enumerate(
            ["","TOTAL","","","","","","","","",
             totals.get("mtm",""),   totals.get("pnl",""),
             totals.get("Delta",""), totals.get("Gamma",""),
             totals.get("Vega",""),  totals.get("Theta",""),
             totals.get("Rho",""),   totals.get("margin","")], 1
        ):
            c = ws4.cell(row=tr, column=ci, value=val)
            c.font = Font(bold=True, color="4E9EFF")
            c.fill = SUB_FILL

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="option_pricing_report.xlsx",
    )



# ─────────────────────────────────────────────────────────
# API: Snowball (Autocallable) Pricing — Monte Carlo
# ─────────────────────────────────────────────────────────

@app.route("/api/snowball", methods=["POST"])
def snowball():
    """
    Price a Snowball (Autocallable) note using path-based Monte Carlo.
    Monthly observation: if S >= KO barrier → knock out, pay coupon accrued.
    At any point if S <= KI barrier → knocked in (becomes vanilla short put).
    If no event → pay full coupon at maturity.

    Request JSON:
        S, sigma, r, q  — standard params
        T               — maturity in years (e.g. 1.0)
        KO_pct          — knock-out level as pct of S0 (e.g. 1.05 = 105%)
        KI_pct          — knock-in  level as pct of S0 (e.g. 0.75 = 75%)
        coupon_pa       — annual coupon rate (e.g. 0.15 = 15%)
        obs_freq        — observations per year (12 = monthly, 4 = quarterly)
        spread          — bid-ask half-spread as fraction of mid (e.g. 0.02 = 2%)
        n_paths         — Monte Carlo paths (default 50000)
    """
    try:
        d        = request.get_json(force=True)
        S0       = float(d["S"])
        sigma    = float(d["sigma"])
        r        = float(d["r"])
        q        = float(d.get("q", 0.0))
        T        = float(d["T"])
        KO_pct   = float(d.get("KO_pct", 1.05))
        KI_pct   = float(d.get("KI_pct", 0.75))
        coupon   = float(d.get("coupon_pa", 0.15))
        freq     = int(d.get("obs_freq", 12))
        spread   = float(d.get("spread", 0.02))
        n        = min(int(d.get("n_paths", 50000)), 200000)
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": f"Invalid parameters: {e}"}), 400

    KO = S0 * KO_pct
    KI = S0 * KI_pct
    dt = 1.0 / freq
    steps = int(round(T * freq))
    np.random.seed(42)

    # Simulate all paths at observation dates
    Z      = np.random.standard_normal((n, steps))
    log_r  = (r - q - 0.5 * sigma ** 2) * dt
    log_s  = sigma * np.sqrt(dt)
    log_ret = log_r + log_s * Z
    # S_paths[i, t] = spot at obs t for path i
    S_paths = S0 * np.exp(np.cumsum(log_ret, axis=1))

    payoffs = np.zeros(n)
    settled = np.zeros(n, dtype=bool)
    ki_hit  = np.zeros(n, dtype=bool)

    for t in range(steps):
        t_yr = (t + 1) * dt
        # Check knock-in (any path that ever goes below KI)
        ki_hit |= S_paths[:, t] <= KI

        # Check knock-out for unsettled paths
        ko_mask = (~settled) & (S_paths[:, t] >= KO)
        if np.any(ko_mask):
            # Knock out: receive coupon accrued to this observation
            payoffs[ko_mask] = np.exp(-r * t_yr) * coupon * t_yr
            settled[ko_mask] = True

    # Paths not knocked out by maturity
    not_settled = ~settled
    if np.any(not_settled):
        final_S = S_paths[not_settled, -1]
        ki_final = ki_hit[not_settled]

        p = np.zeros(np.sum(not_settled))
        # No KI hit → receive full coupon
        p[~ki_final] = np.exp(-r * T) * coupon * T
        # KI hit → short put payoff (loss if S < S0)
        p[ki_final]  = np.exp(-r * T) * np.minimum(final_S[ki_final] / S0 - 1.0, 0.0)
        payoffs[not_settled] = p

    mid   = round(float(np.mean(payoffs)), 6)
    std   = round(float(np.std(payoffs) / np.sqrt(n)), 6)
    ci95  = round(1.96 * float(np.std(payoffs) / np.sqrt(n)), 6)
    bid   = round(mid - abs(mid) * spread, 6)
    ask   = round(mid + abs(mid) * spread, 6)
    ko_prob = round(float(np.mean(settled)), 4)
    ki_prob = round(float(np.mean(ki_hit)), 4)

    return jsonify({
        "mid": mid, "bid": bid, "ask": ask,
        "std_err": std, "ci95": ci95,
        "ko_prob": ko_prob, "ki_prob": ki_prob,
        "n_paths": n,
        "params": {"KO": KO, "KI": KI, "coupon_pa": coupon,
                   "T": T, "freq": freq}
    })


# ─────────────────────────────────────────────────────────
# API: Shark Fin (Barrier Option) Pricing — Closed Form
# ─────────────────────────────────────────────────────────

@app.route("/api/shark_fin", methods=["POST"])
def shark_fin():
    """
    Price barrier options (Up-and-Out Call / Down-and-Out Put) analytically.
    Uses the Rubinstein-Reiner / Haug (1998) closed-form formula.

    For Up-and-Out Call (UOC): η = -1, φ = 1, requires H > K
    For Down-and-Out Put (DOP): η = +1, φ = -1, requires H < K

    Correct formula: price = A - B - C + D  (with η-signed N() arguments)

    Request JSON:
        flag    — 'c' (Up-and-Out Call) or 'p' (Down-and-Out Put)
        S, K, T, r, sigma, q — standard BSM params
        H       — barrier level
        rebate  — cash rebate paid if barrier is hit (default 0)
        spread  — bid-ask half-spread fraction (default 0.02)
    """
    try:
        d       = request.get_json(force=True)
        flag    = str(d.get("flag", "c")).lower()
        S       = float(d["S"])
        K       = float(d["K"])
        T       = float(d["T"])
        r       = float(d["r"])
        sigma   = float(d["sigma"])
        q       = float(d.get("q", 0.0))
        H       = float(d["H"])
        rebate  = float(d.get("rebate", 0.0))
        spread  = float(d.get("spread", 0.02))
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": f"Invalid parameters: {e}"}), 400

    if T <= 0 or sigma <= 0:
        return jsonify({"error": "T and sigma must be positive"}), 400

    from scipy.stats import norm

    # ── Haug (1998) closed-form for continuous barrier options ──────────────
    phi = 1.0 if flag == 'c' else -1.0   # +1 call, -1 put
    eta = -1.0 if flag == 'c' else 1.0   # -1 up barrier, +1 down barrier

    if flag == 'c' and H <= S:
        return jsonify({"error": "For Up-and-Out Call, barrier H must be above current spot S"}), 400
    if flag == 'p' and H >= S:
        return jsonify({"error": "For Down-and-Out Put, barrier H must be below current spot S"}), 400
    if flag == 'c' and H <= K:
        return jsonify({"error": "For Up-and-Out Call, H must be > K (otherwise always knocked out)"}), 400
    if flag == 'p' and H >= K:
        return jsonify({"error": "For Down-and-Out Put, H must be < K (otherwise always knocked out)"}), 400

    try:
        sT  = sigma * np.sqrt(T)
        # λ = (r - q + σ²/2) / σ²  [Haug convention]
        lam = (r - q + 0.5 * sigma**2) / sigma**2
        # μ = (r - q - σ²/2) / σ²  = λ - 1
        mu  = lam - 1.0

        eqT = np.exp(-q * T)
        erT = np.exp(-r * T)
        hS  = H / S           # H/S ratio

        # BSM argument definitions (Haug p.66)
        x1 = np.log(S / K) / sT + lam * sT
        x2 = np.log(S / H) / sT + lam * sT
        y1 = np.log(H**2 / (S * K)) / sT + lam * sT
        y2 = np.log(H / S) / sT + lam * sT

        def _A(x):
            """Standard BSM-style term using φ sign."""
            return (phi * S * eqT * norm.cdf(phi * x)
                    - phi * K * erT * norm.cdf(phi * x - phi * sT))

        def _B(y):
            """Reflection term using η sign on N() arguments."""
            return (phi * S * eqT * hS**(2 * lam) * norm.cdf(eta * y)
                    - phi * K * erT * hS**(2 * lam - 2) * norm.cdf(eta * y - eta * sT))

        A = _A(x1)   # = vanilla option price
        B = _A(x2)   # barrier-adjusted BSM term
        C = _B(y1)   # reflection term with y1
        D = _B(y2)   # reflection term with y2

        barrier_price = float(A - B - C + D)

        # Rebate present value (paid at barrier hit time)
        if rebate > 0:
            z  = np.log(H / S) / sT + lam * sT
            r1 = hS ** (mu + lam)
            r2 = hS ** (mu - lam)
            rebate_pv = float(rebate * (
                r1 * norm.cdf(eta * z) +
                r2 * norm.cdf(eta * z - 2 * eta * lam * sT)
            ))
        else:
            rebate_pv = 0.0

        # Vanilla price for comparison
        vanilla = float(A)

    except Exception as e:
        return jsonify({"error": f"Pricing error: {e}"}), 500

    mid = round(max(0.0, barrier_price + rebate_pv), 6)
    bid = round(mid * (1 - spread), 6)
    ask = round(mid * (1 + spread), 6)

    return jsonify({
        "mid": mid, "bid": bid, "ask": ask,
        "vanilla_price": round(vanilla, 6),
        "barrier_discount": round(max(0.0, vanilla - barrier_price), 6),
        "rebate_pv": round(rebate_pv, 6),
        "params": {"H": H, "rebate": rebate, "flag": flag}
    })


# ─────────────────────────────────────────────────────────
# API: OTC Forward Pricing
# ─────────────────────────────────────────────────────────

@app.route("/api/forward", methods=["POST"])
def otc_forward():
    """
    Price an OTC Forward contract.  F = S * exp((r - q) * T)

    Request JSON:
        S, T, r, q  — standard params
        direction   — 'long' or 'short'
        notional    — contract notional (default 1)
        spread      — bid-ask half-spread as absolute points on F (default 0)
    """
    try:
        d          = request.get_json(force=True)
        S          = float(d["S"])
        T          = float(d["T"])
        r          = float(d["r"])
        q          = float(d.get("q", 0.0))
        direction  = str(d.get("direction", "long")).lower()
        notional   = float(d.get("notional", 1.0))
        spread_pts = float(d.get("spread", 0.0))
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": f"Invalid parameters: {e}"}), 400

    F             = S * np.exp((r - q) * T)
    cost_of_carry = S * (np.exp((r - q) * T) - 1.0)
    funding_cost  = S * (np.exp(r * T) - 1.0)
    div_income    = S * (np.exp(q * T) - 1.0)

    mid = round(float(F), 6)
    bid = round(mid - spread_pts, 6)
    ask = round(mid + spread_pts, 6)

    sign = 1 if direction == "long" else -1
    mtm_per_unit = sign * (F - S)  # theoretical P&L vs spot

    return jsonify({
        "forward_price": mid,
        "bid": bid, "ask": ask,
        "cost_of_carry":   round(float(cost_of_carry), 4),
        "funding_cost":    round(float(funding_cost), 4),
        "dividend_income": round(float(div_income), 4),
        "mtm_per_unit":    round(float(mtm_per_unit), 4),
        "notional_mtm":    round(float(mtm_per_unit * notional), 4),
        "params": {"S": S, "T": T, "r": r, "q": q,
                   "direction": direction, "notional": notional}
    })


# ─────────────────────────────────────────────────────────
# API: Interest Rate Swap (IRS) — Simplified DCF
# ─────────────────────────────────────────────────────────

@app.route("/api/irs", methods=["POST"])
def irs():
    """
    Price a plain vanilla Interest Rate Swap (IRS) using simplified flat-curve DCF.

    Pay-fixed / receive-float convention (NPV from perspective of fixed payer).
    NPV = PV(Float leg) - PV(Fixed leg)

    Request JSON:
        notional      — notional principal
        fixed_rate    — annual fixed rate (e.g. 0.035 = 3.5%)
        float_rate    — current floating rate / par rate (e.g. 0.025 = SHIBOR)
        T             — maturity in years
        freq          — payment frequency per year (1=annual, 2=semi, 4=quarterly)
        r_disc        — discount rate (defaults to float_rate)
        spread        — bid-ask half-spread as absolute NPV (default 0)
    """
    try:
        d           = request.get_json(force=True)
        notional    = float(d.get("notional", 1_000_000))
        fixed_rate  = float(d["fixed_rate"])
        float_rate  = float(d["float_rate"])
        T           = float(d["T"])
        freq        = int(d.get("freq", 4))
        r_disc      = float(d.get("r_disc", float_rate))
        spread      = float(d.get("spread", 0.0))
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": f"Invalid parameters: {e}"}), 400

    dt = 1.0 / freq
    periods = int(round(T * freq))
    if periods < 1:
        return jsonify({"error": "Too short: T * freq < 1"}), 400

    fixed_cf  = fixed_rate * dt * notional
    float_cf  = float_rate * dt * notional

    cashflows = []
    pv_fixed = pv_float = 0.0
    for i in range(1, periods + 1):
        t_i    = i * dt
        df     = np.exp(-r_disc * t_i)
        # Add principal repayment to last period
        extra  = notional if i == periods else 0.0
        fix_pmt = fixed_cf + extra
        flt_pmt = float_cf + extra
        pv_f   = fix_pmt * df
        pv_fl  = flt_pmt * df
        pv_fixed  += pv_f
        pv_float  += pv_fl
        cashflows.append({
            "period":     i,
            "t":          round(t_i, 4),
            "df":         round(float(df), 6),
            "fixed_pmt":  round(fix_pmt, 2),
            "float_pmt":  round(flt_pmt, 2),
            "pv_fixed":   round(pv_f, 2),
            "pv_float":   round(pv_fl, 2),
            "net_pmt":    round(flt_pmt - fix_pmt, 2),
        })

    npv  = round(float(pv_float - pv_fixed), 2)
    bid  = round(npv - spread, 2)
    ask  = round(npv + spread, 2)
    dv01 = round(float(notional * dt * np.exp(-r_disc * T) * 0.0001), 2)  # approx DV01

    return jsonify({
        "npv": npv, "bid": bid, "ask": ask,
        "pv_fixed":  round(float(pv_fixed), 2),
        "pv_float":  round(float(pv_float), 2),
        "dv01":      dv01,
        "cashflows": cashflows,
        "params": {
            "notional": notional, "fixed_rate": fixed_rate,
            "float_rate": float_rate, "T": T, "freq": freq
        }
    })




if __name__ == "__main__":
    print("=" * 60)
    print("  Option Pricing Tool -- Web Application")
    print("  Open in browser: http://127.0.0.1:5000")
    print("=" * 60)
    app.run(debug=True, port=5000)
